import pandas as pd
import streamlit as st
from io import BytesIO
import re
from openpyxl.styles import PatternFill
import os

# App title
st.title("Fuckin' Awesome File Convertor")

# Define the path to the shipping legend
shipping_legend_path = "project-folder/data/default_shipping_legend.xlsx"

# Path for Blocked Brands file
blocked_brands_path = "project-folder/data/Blocked_Brands.xlsx"

# Ensure the Blocked Brands file exists and contains the correct worksheet
if not os.path.exists(blocked_brands_path):
    # Create a new Blocked Brands file with the correct structure if it doesn't exist
    pd.DataFrame(columns=["Blocked Brands"]).to_excel(blocked_brands_path, index=False, sheet_name="Blocked_Brands")
else:
    # Check if the worksheet exists; create it if it doesn't
    try:
        excel_data = pd.ExcelFile(blocked_brands_path)
        if "Blocked_Brands" not in excel_data.sheet_names:
            pd.DataFrame(columns=["Blocked Brands"]).to_excel(blocked_brands_path, index=False, sheet_name="Blocked_Brands")
    except Exception as e:
        st.error(f"Error initializing Blocked Brands file: {e}")

# Sidebar form to add a blocked brand
st.sidebar.header("Manage Blocked Brands")
with st.sidebar.form("Add Blocked Brands"):
    new_brand = st.text_input("Enter the brand to block")
    submit_button = st.form_submit_button("Add Brand")

    if submit_button:
        try:
            # Load the existing Blocked Brands list
            blocked_brands = pd.read_excel(blocked_brands_path, sheet_name="Blocked_Brands")
            
            # Check if the brand is already blocked
            if new_brand and new_brand.strip() not in blocked_brands["Blocked Brands"].values:
                # Append the new brand
                new_brand_df = pd.DataFrame([{"Blocked Brands": new_brand.strip()}])
                blocked_brands = pd.concat([blocked_brands, new_brand_df], ignore_index=True)
                
                # Save back to the Excel file
                with pd.ExcelWriter(blocked_brands_path, engine="openpyxl", mode="w") as writer:
                    blocked_brands.to_excel(writer, index=False, sheet_name="Blocked_Brands")
                
                st.sidebar.success(f"Brand '{new_brand}' has been added to the blocked list.")
            elif new_brand.strip() in blocked_brands["Blocked Brands"].values:
                st.sidebar.warning(f"The brand '{new_brand}' is already in the blocked list.")
            else:
                st.sidebar.warning("Please enter a valid brand name.")
        except Exception as e:
            st.sidebar.error(f"Error updating blocked brands: {e}")

# Bulk Upload for Blocked Brands
st.sidebar.subheader("Bulk Upload Blocked Brands")
bulk_file = st.sidebar.file_uploader("Upload an Excel file with Blocked Brands", type=["xlsx"])

if bulk_file:
    try:
        bulk_brands = pd.read_excel(bulk_file)
        if "Blocked Brands" not in bulk_brands.columns:
            st.sidebar.error("The uploaded file must contain a 'Blocked Brands' column.")
        else:
            blocked_brands = pd.read_excel(blocked_brands_path, sheet_name="Blocked_Brands")
            # Combine the current and new brands, avoiding duplicates
            updated_brands = pd.concat([blocked_brands, bulk_brands]).drop_duplicates(subset=["Blocked Brands"], ignore_index=True)
            
            # Save back to the Blocked Brands file
            with pd.ExcelWriter(blocked_brands_path, engine="openpyxl", mode="w") as writer:
                updated_brands.to_excel(writer, index=False, sheet_name="Blocked_Brands")
            
            st.sidebar.success("Blocked Brands have been updated successfully.")
    except Exception as e:
        st.sidebar.error(f"Error processing bulk upload: {e}")

# Display and Manage Blocked Brands
try:
    # Load the Blocked Brands file
    blocked_brands = pd.read_excel(blocked_brands_path, sheet_name="Blocked_Brands")

    # Ensure proper serial numbers without duplication
    blocked_brands.reset_index(drop=True, inplace=True)  # Drop existing index
    blocked_brands.insert(0, "S.No", range(1, len(blocked_brands) + 1))  # Add "S.No" as the first column

    # Display only the S.No and Blocked Brands columns
    st.sidebar.subheader("Blocked Brands")
    st.sidebar.write(blocked_brands)

    # Provide a download button for the Blocked Brands list
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        blocked_brands.to_excel(writer, index=False, sheet_name="Blocked_Brands")
    buffer.seek(0)

    st.sidebar.download_button(
        label="Download Blocked Brands",
        data=buffer.getvalue(),
        file_name="Blocked_Brands.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.sidebar.error(f"Error loading blocked brands: {e}")


    # Provide a download button for the Blocked Brands list
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        blocked_brands[["Blocked Brands"]].to_excel(writer, index=False, sheet_name="Blocked_Brands")
    buffer.seek(0)

    st.sidebar.download_button(
        label="Download Blocked Brands",
        data=buffer.getvalue(),
        file_name="Blocked_Brands.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.sidebar.error(f"Error loading blocked brands: {e}")


    # Provide a download button for the Blocked Brands list
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        blocked_brands[["Blocked Brands"]].to_excel(writer, index=False, sheet_name="Blocked_Brands")
    buffer.seek(0)

    st.sidebar.download_button(
        label="Download Blocked Brands",
        data=buffer.getvalue(),
        file_name="Blocked_Brands.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.sidebar.error(f"Error loading blocked brands: {e}")


    # Provide a download button for the blocked brands file
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        blocked_brands.to_excel(writer, index=False, sheet_name="Blocked_Brands")
    buffer.seek(0)

    st.sidebar.download_button(
        label="Download Blocked Brands",
        data=buffer.getvalue(),
        file_name="Blocked_Brands.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.sidebar.error(f"Error loading blocked brands: {e}")


    
# Step 1: File uploader
st.header("Upload Excel Files")
uploaded_files = st.file_uploader("Upload one or more Excel files", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    all_data = []

    # Step 2: Process each uploaded file
    for uploaded_file in uploaded_files:
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            for sheet_name in excel_file.sheet_names:
                sheet_data = pd.read_excel(uploaded_file, sheet_name=sheet_name, usecols="B,E,G,H,I")
                all_data.append(sheet_data)
        except Exception as e:
            st.error(f"Error reading file {uploaded_file.name}: {e}")

    if all_data:
        # Step 3: Combine all sheets into one DataFrame
        combined_df = pd.concat(all_data, ignore_index=True)
        st.write("### Combined Data Preview (Before Renaming)")
        st.dataframe(combined_df)

        if uploaded_files:
            all_data = []

    # Process each uploaded file
    for uploaded_file in uploaded_files:
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            for sheet_name in excel_file.sheet_names:
                sheet_data = pd.read_excel(uploaded_file, sheet_name=sheet_name, usecols="B,E,G,H,I")
                all_data.append(sheet_data)
        except Exception as e:
            st.error(f"Error reading file {uploaded_file.name}: {e}")

        # Step 3.1: Add HANDLING COST column
        st.write("### Adding HANDLING COST Column")
        combined_df["HANDLING COST"] = 0.75
        st.success("HANDLING COST column added with default value 0.75.")

        # Step 4: Standardize and Rename Columns
        st.write("### Renaming Columns")
        combined_df.columns = combined_df.columns.str.strip()  # Strip column headers of extra spaces
        column_mapping = {
            "Product Details": "TITLE",
            "Brand": "BRAND",
            "Product ID": "SKU",
            "UPC Code": "UPC/ISBN",
            "Price": "COST_PRICE"
        }
        combined_df.rename(columns=column_mapping, inplace=True)

        if "COST_PRICE" not in combined_df.columns:
            st.error("COST_PRICE column is missing. Ensure the input file has a 'Price' column.")
        else:
            st.success("Columns renamed successfully.")

        # Step 4.1: Format SKU column to remove commas and ensure it is displayed as a string
        if "SKU" in combined_df.columns:
            combined_df["SKU"] = combined_df["SKU"].astype(str).str.replace(",", "").str.strip()
            st.success("SKU column formatted to remove commas.")

        # Step 5: Clean TITLE column
        if "TITLE" in combined_df.columns:
            combined_df["TITLE"] = (
                combined_df["TITLE"]
                .str.replace(r"\(W\+\)", "", regex=True)
                .str.replace(r"\(SP\)", "", regex=True)
                .str.replace(r"\(P\)", "", regex=True)
                .str.strip()
            )
            st.success("Unwanted patterns removed from TITLE column.")

        # Step 6: Format UPC/ISBN column
        st.write("### Formatting UPC/ISBN Column")
        if "UPC/ISBN" in combined_df.columns:
            combined_df["UPC/ISBN"] = (
            combined_df["UPC/ISBN"]
            .apply(lambda x: str(int(float(x))) if pd.notnull(x) else "")  # Convert to integer, then string
            .str.zfill(12)  # Add leading zeros to ensure 12 digits
        )
        st.success("UPC/ISBN column formatted to have a minimum of 12 digits as a string.")


        # Step 7: Format COST_PRICE column
        st.write("### Formatting COST_PRICE Column")
        if "COST_PRICE" in combined_df.columns:
            combined_df["COST_PRICE"] = (
                combined_df["COST_PRICE"]
                .astype(str)
                .str.replace(r"[$,]", "", regex=True)  # Remove currency symbols and commas
                .astype(float)
                .round(2)
            )
            st.success("COST_PRICE column formatted to numeric with two decimal places.")

        # Step 8: Add QUANTITY and ITEM LOCATION columns
        combined_df["QUANTITY"] = 1
        combined_df["ITEM LOCATION"] = "WALMART"

        # Step 9: Add ITEM WEIGHT (pounds) column
        st.write("### Adding ITEM WEIGHT (pounds) Column")

        # Function to extract weight and handle pack sizes
        def extract_weight_with_packs(title):
            """
            Extract the weight and account for pack size in the TITLE.
            """
            try:
                # Extract the weight (e.g., "8 oz", "10 fl oz", etc.)
                match_weight = re.search(r"(\d+(\.\d+)?)\s*(?:oz|ounces|ounce|fl. oz.|fluid ounce|fl oz|fluid ounces)", title, re.IGNORECASE)
                single_unit_weight = float(match_weight.group(1)) if match_weight else None

                # Extract the pack size (e.g., "2 pack", "pack of 3", etc.)
                match_pack = re.search(r"(?:\b(\d+)\s*pack\b|\bpack of\s*(\d+))", title, re.IGNORECASE)
                pack_size = int(match_pack.group(1) or match_pack.group(2)) if match_pack else 1  # Default to 1 if no pack

                if single_unit_weight is not None:
                    # Add 6 oz or 10 oz based on the unit type and calculate the total weight
                    if match_weight and "fl oz" in match_weight.group(0).lower():
                        single_unit_weight += 10  # Add 10 oz for "fl oz"
                    else:
                        single_unit_weight += 6  # Add 6 oz for "oz" or "ounces"

                    # Calculate total weight for the pack and convert to pounds
                    total_weight = (single_unit_weight * pack_size) / 16  # Convert oz to pounds
                    return round(total_weight, 2)

            except Exception as e:
                # Log the error for debugging
                st.error(f"Error processing title '{title}': {e}")

            # Return None if no weight or pack size is found
            return None

        # Apply the function to the TITLE column
        if "TITLE" in combined_df.columns:
            combined_df["ITEM WEIGHT (pounds)"] = combined_df["TITLE"].apply(
                lambda x: extract_weight_with_packs(x) if isinstance(x, str) else None
            )
            st.success("ITEM WEIGHT (pounds) column updated to account for pack sizes.")

        # Step 9.1: Highlight rows with missing weights
        st.write("### Highlighting Rows with Missing ITEM WEIGHT (pounds)")


        # Step 10: Add SHIPPING COST column
        st.write("### Adding SHIPPING COST Column")

        shipping_legend_path = "project-folder/data/default_shipping_legend.xlsx"  # Relative path in the repository

        if not os.path.exists(shipping_legend_path):
            st.error(f"The shipping legend file does not exist at the specified path: {shipping_legend_path}")
        else:
            try:
                shipping_legend = pd.read_excel(shipping_legend_path, engine="openpyxl")
                st.success("Shipping legend file loaded successfully.")
            except Exception as e:
                st.error(f"Error reading shipping legend file: {e}")
                shipping_legend = None

        if shipping_legend is not None and {"Weight Range Min (lb)", "Weight Range Max (lb)", "SHIPPING COST"}.issubset(shipping_legend.columns):
            def calculate_shipping_cost(weight, legend):
                if pd.isnull(weight):
                    return None
                for _, row in legend.iterrows():
                    if row["Weight Range Min (lb)"] <= weight <= row["Weight Range Max (lb)"]:
                        return row["SHIPPING COST"]
                return None

            combined_df["SHIPPING COST"] = combined_df["ITEM WEIGHT (pounds)"].apply(
                lambda w: calculate_shipping_cost(w, shipping_legend)
            )
        else:
            st.error("Shipping legend file is missing required columns: 'Weight Range Min (lb)', 'Weight Range Max (lb)', 'SHIPPING COST'.")

        # Step 10.1: Add RETAIL PRICE column
        if all(col in combined_df.columns for col in ["COST_PRICE", "SHIPPING COST", "HANDLING COST"]):
            combined_df["RETAIL PRICE"] = combined_df.apply(
                lambda row: round(
                    (row["COST_PRICE"] + row["SHIPPING COST"] + row["HANDLING COST"]) * 1.35, 2
                ) if not (pd.isnull(row["COST_PRICE"]) or pd.isnull(row["SHIPPING COST"]) or pd.isnull(row["HANDLING COST"])) else None,
                axis=1
            )

        # Step 10.2: Add MIN PRICE and MAX PRICE columns
        if all(col in combined_df.columns for col in ["SHIPPING COST", "ITEM WEIGHT (pounds)", "RETAIL PRICE"]):
            combined_df["MIN PRICE"] = combined_df["RETAIL PRICE"]
            combined_df["MAX PRICE"] = combined_df["RETAIL PRICE"].apply(lambda x: round(x * 1.35, 2) if x is not None else None)

       

        # Step 11: Remove duplicate rows
        combined_df.drop_duplicates(inplace=True)



    # Step 11.1: Calculate and Display Metrics
    st.write("### Metrics Summary")

    # Total number of listings in the input files
    total_input_listings = len(pd.concat(all_data, ignore_index=True))

    # Total number of listings in the output file
    total_output_listings = len(combined_df)

    # Total duplicates removed
    total_duplicates_removed = total_input_listings - total_output_listings

    # Total listings with no weights
    listings_no_weights = combined_df["ITEM WEIGHT (pounds)"].isnull().sum()    
        
    # Step X.1: Update metrics
    removed_blocked_count = len(removed_rows) if "removed_rows" in locals() else 0

    # Display the metrics
    st.markdown(f"""
    - **Total Listings in Input Files:** {total_input_listings}
    - **Total Listings in Output File:** {total_output_listings}
    - **Total Duplicates Removed:** {total_duplicates_removed}
    - **Listings with No Weights (Red Highlighted Rows):** {listings_no_weights}
    """)


    # Step 12: Export final DataFrame with Conditional Formatting
    st.write("### Download Consolidated File")

    # Step 12.1: Move rows with missing weights to the end
    st.write("### Reordering Rows with Missing Weights")
    combined_df['Missing Weight'] = combined_df['ITEM WEIGHT (pounds)'].isnull()
    combined_df = combined_df.sort_values(by='Missing Weight', ascending=True).drop(columns=['Missing Weight'])
    st.success("Rows with missing weights have been moved to the bottom.")

    # Step 12.2: Define a styling function for highlighting rows
    def highlight_missing_weights(row):
        if pd.isnull(row["ITEM WEIGHT (pounds)"]):
            return ["background-color: #FFCCCC"] * len(row)
        return [""] * len(row)

    # Step 12.4: Format numeric columns to 2 decimal places
    numeric_columns = [
        "COST_PRICE",
        "HANDLING COST",
        "ITEM WEIGHT (pounds)",
        "SHIPPING COST",
        "RETAIL PRICE",
        "MIN PRICE",
        "MAX PRICE",
    ]

    # Apply formatting for numeric columns
    styled_df = (
        combined_df.style.apply(highlight_missing_weights, axis=1)
        .format({col: "{:.2f}" for col in numeric_columns})
    )

    # Display the styled DataFrame with formatting
    st.write("### Updated Final Data Preview with Highlights and Formatting")
    st.dataframe(styled_df)

if st.button("Export to Excel"):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Write the main DataFrame to the first sheet
        combined_df.to_excel(writer, index=False, sheet_name="Consolidated Data")

        # Embed the shipping legend as a separate sheet
        if shipping_legend is not None:
            shipping_legend.to_excel(writer, index=False, sheet_name="ShippingLegend")

        # Access the "Consolidated Data" worksheet
        worksheet = writer.sheets["Consolidated Data"]

        # Define a red fill style for missing weights
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Iterate over rows and add formulas for missing weights
        for row_index, weight in enumerate(combined_df["ITEM WEIGHT (pounds)"], start=2):  # Start from row 2 (Excel row 2)
            if pd.isnull(weight):
                # Highlight the row in red
                for col_index in range(1, len(combined_df.columns) + 1):
                    worksheet.cell(row=row_index, column=col_index).fill = red_fill

                # Add formulas to the relevant columns
                worksheet.cell(row=row_index, column=10).value = f"=IF(I{row_index}<>\"\", ROUND(VLOOKUP(I{row_index}, ShippingLegend!A:C, 3, TRUE), 2), \"\")" # SHIPPING COST formula
                worksheet.cell(row=row_index, column=11).value = (f"=IF(AND(E{row_index}<>\"\", F{row_index}<>\"\", J{row_index}<>\"\"), "f"ROUND((E{row_index}+F{row_index}+J{row_index})*1.35, 2), \"\")")  # RETAIL PRICE formula                
                worksheet.cell(row=row_index, column=12).value = f"=K{row_index}"  # MIN PRICE formula
                worksheet.cell(row=row_index, column=13).value = f"=IF(L{row_index}<>\"\", ROUND(L{row_index}*1.35, 2),\"\")"  # MAX PRICE formula

    # Save the updated Excel file
    buffer.seek(0)
    st.download_button(
        label="Download Excel File",
        data=buffer.getvalue(),
        file_name="Consolidated_Data_with_Embedded_Legend.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if combined_df.empty:
        st.info("Upload one or more Excel files to get started.")
    else:
        st.success("The output file is ready for download.")
